from bs4 import BeautifulSoup as bs
import requests as req
from openpyxl import load_workbook, Workbook

url = "http://www.basketball-reference.com/leagues"
user_agent = {'User-agent': "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36"}

wb = Workbook()

rep = req.get(url, headers=user_agent)
soup = bs(rep.text, "html.parser")
tbs = soup.find_all("table")
trs = tbs[0].find_all("tr")

ws = wb.create_sheet()

for i, tr in enumerate(trs):
	for j, td in enumerate(tr.findChildren(["th", "td"])):
		ws.cell(row=i+1, column=j+1, value=td.get_text())
		#print(td.get_text(), end=' ')
	#print(end='\n')

wb.save("test.xlsx")


