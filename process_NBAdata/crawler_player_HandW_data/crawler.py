from bs4 import BeautifulSoup
import urllib.request
from openpyxl import load_workbook, Workbook

a = 97
wb = Workbook()
for num in range(26):
	print("num: ", a)
	url = "http://www.basketball-reference.com/players/"+chr(a)
	try:
		response = urllib.request.urlopen(url)
		html = response.read()
		soup = BeautifulSoup(html)

		tbs = soup.find_all("table")
		trs = tbs[0].find_all('tr')

		ws = wb.create_sheet(chr(a))

		for i, tr in enumerate(trs):
		    for j, td in enumerate(tr.findChildren(['th', 'td'])):
		    	# print(td.get_text(), end = " ")
		    	ws.cell(column = j+1, row = i+1, value=str(td.get_text()))

		a +=  1
	except:
		a +=  1

wb.save("/Users/mars/Downloads/test.xlsx")
